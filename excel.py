import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import Workbook


def extract_and_manipulate_data(year):
    # Create a Tkinter root window and hide it
    root = tk.Tk()
    root.withdraw()

    # Open a file dialog to select the Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])

    if file_path:
        try:
            # Load the Excel file into a pandas DataFrame
            data = pd.read_excel(file_path)

            # Filter "Receive Date" by selecting the specified year
            data['Receive Date'] = pd.to_datetime(data['Receive Date'])
            data_year = data[data['Receive Date'].dt.year == year]

            # Create dictionaries to store the counts of serial numbers in Jan - Jul and Aug - Dec
            jan_to_jul_counts = {}
            aug_to_dec_counts = {}

            # Iterate through the DataFrame and count serial numbers within the specified periods
            for index, row in data_year.iterrows():
                receive_month = row['Receive Date'].month
                serial_no = row['Serial No.']
                if 1 <= receive_month <= 7:
                    if serial_no not in jan_to_jul_counts:
                        jan_to_jul_counts[serial_no] = 0
                    jan_to_jul_counts[serial_no] += 1
                elif 8 <= receive_month <= 12:
                    if serial_no not in aug_to_dec_counts:
                        aug_to_dec_counts[serial_no] = 0
                    aug_to_dec_counts[serial_no] += 1

            # Find and extract duplicates from "Serial No."
            duplicates = data_year[data_year.duplicated(subset=['Serial No.'], keep=False)]
            unique_data = data_year[~data_year.duplicated(subset=['Serial No.'], keep=False)]

            # Save the extracted unique data and duplicates to separate sheets in a new Excel file
            save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[("Excel Files", "*.xlsx")])

            if save_path:
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    unique_data.to_excel(writer, sheet_name='Unique Data', index=False)

                    # Group duplicates by 'Serial No.' and create a new DataFrame for formatting
                    duplicates_grouped = duplicates.groupby('Serial No.', as_index=False)['Receive Date'].agg(list)

                    formatted_data = []
                    for row in duplicates_grouped.itertuples(index=False):
                        serial_no = row[0]
                        receive_dates = row[1]
                        total_count = len(receive_dates)  # Calculate total count of Receive Dates
                        formatted_data.append(
                            {'Serial No.': serial_no, 'Receive Date': receive_dates[0], 'Total Count': total_count})
                        for date in receive_dates[1:]:
                            formatted_data.append({'Serial No.': '', 'Receive Date': date, 'Total Count': ''})

                    formatted_df = pd.DataFrame(formatted_data)
                    formatted_df.to_excel(writer, sheet_name='Duplicates', index=False)

                    # Get the worksheet
                    worksheet = writer.sheets['Duplicates']

                    # Define yellow fill color
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                    # Apply formatting based on Total Count
                    for row_idx, row in enumerate(
                            worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3)):
                        total_count = row[0].value
                        if total_count and total_count > 1:
                            for cell in row:
                                cell.fill = yellow_fill

                    # Add "Jan - Jul" and "Aug - Dec" columns to the "Duplicates" sheet
                    worksheet['D1'] = "Jan - Jul"
                    worksheet['E1'] = "Aug - Dec"

                    # Populate "Jan - Jul" and "Aug - Dec" columns in the "Duplicates" sheet based on the counts
                    for row_idx, row in enumerate(
                            worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1)):
                        serial_no = row[0].value
                        jan_to_jul_count = jan_to_jul_counts.get(serial_no, 0)
                        aug_to_dec_count = aug_to_dec_counts.get(serial_no, 0)
                        worksheet.cell(row=row_idx + 2, column=4, value=jan_to_jul_count)
                        worksheet.cell(row=row_idx + 2, column=5, value=aug_to_dec_count)

                print("Data extracted, manipulated, and saved successfully!")

        except Exception as e:
            print("An error occurred:", e)


# Call the function for different years
extract_and_manipulate_data(2020)
extract_and_manipulate_data(2021)
extract_and_manipulate_data(2022)
extract_and_manipulate_data(2023)
